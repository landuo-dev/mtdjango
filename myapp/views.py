import asyncio
import os
import re
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Color, Font
from django.core.files.storage import FileSystemStorage
import pandas as pd
import sys
import redis

from django_redis import get_redis_connection
from pymongo import MongoClient
from io import BytesIO
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.core.cache import cache

from Custom_features.店铺信息入库 import get_product as set_database_jx01

from Custom_features.协程批量修改夹心 import main02 as updata_Database_jx02
from Custom_features.批量改夹心 import main02 as set_jx
# from Custom_features.批量修改夹心 import main1 as updata_Database_jx02
from Custom_features.折扣信息入库补充 import main1 as set_database_jg01, main02 as set_database_jg02
from Custom_features.批量更改商品折扣价 import main as updata_database_jg
from Custom_features.折扣同步2 import main as syncjg
from Custom_features.搬菜单 import move_pro as move_pro
# from Custom_features.月售替换 import rep_pro as rep_pro1
# from Custom_features.月售替换2 import rep_pro as rep_pro2
from Custom_features.月售替换3 import rep_pro as rep_pro3
from Custom_features.名字替换 import rep_name
from Custom_features.auto_reply import auto_rep
from Custom_features.save_ds import save_database as save_ds
from Custom_features.set_jzyx import main as yx
from Custom_features.save_yx import save_jz
from Custom_features.push_ds import push_ds
from Custom_features.delds_rw import delds_rw
from Custom_features.push_json import push_json
from Custom_features.hangye_json import hangye_json
from Custom_features.back_wm import back
from Custom_features.get_back import get_back_log
from Custom_features.get_back_jg import back_zk
from Custom_features.恢复夹心 import Recover
from Custom_features.恢复价格 import Recover_jg
from Custom_features.每日数据 import get_data
from encrypt.JX import JX


# from encrypt.JG import JG


# Create your views here.


def index(request):
    try:
        poi_id = request.session['poi_id']
        client = MongoClient('mongodb://localhost:27017/')
        db = client['other']
        collection = db['wmname']
        wmname = collection.find_one({"id": int(poi_id)})
    except Exception as e:
        return render(request, 'index.html', {"poi_id": '', "num": -1, "aa": -1})

    if wmname:
        name = wmname['poiName']
    else:
        name = ''

    return render(request, 'index.html', {"poi_id": poi_id, "wmname": name, "num": -1, "aa": -1})


# def requires_permission(request):
#     def check1(func):
#         def wrapper(*args, **kwargs):
#             if cache.has_key(request.POST['poi_id']):
#                 return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
#
#             return func(*args, **kwargs)
#         return wrapper
#     return check1


def push_arr():
    client = MongoClient('mongodb://localhost:27017')
    db = client['other']
    collections = db['jt_all']
    arr = [(f"{i['店铺名称']}:{str(int(i['ID号码']))}: {i['店铺对接负责人']}", str(int(i['ID号码']))) for i in collections.find()]
    return arr


def hangye(request, poi_id, startTime=1716134400, endTime=''):
    if poi_id == '':
        poi_id = '11111111111'
    if endTime == '':
        now = datetime.now()
        endTime = int(now.timestamp())

    arr = push_arr()
    print(startTime)
    print(endTime)
    print(poi_id)
    json_Data = hangye_json(poi_id, int(startTime), int(endTime))
    # json_Data['arr'] = arr
    # print(json_Data)
    json_Data["arr"] = arr
    # print(json_Data)
    json_Data['startTime'] = int(startTime)
    json_Data['endTime'] = endTime
    return render(request, 'hangye.html', {"data": json_Data, "poi_id": poi_id})


def push_qd(request, poi_id, startTime=1716134400, endTime=''):
    if poi_id == '':
        poi_id = '10747973'
    arr = push_arr()

    if endTime == '':
        now = datetime.now()
        endTime = int(now.timestamp())

    print(startTime)
    print(endTime)
    print(poi_id)
    json_Data = push_json(poi_id, int(startTime), int(endTime))
    json_Data['arr'] = arr
    # print(json_Data)
    return render(request, 'huatu.html', json_Data)


def get_tupian(request):
    date = request.GET.get('date')
    poi_id = request.GET.get('poi_id')
    client = MongoClient('mongodb://localhost:27017')
    db = client['other']
    collections = db['Essay']
    doc = collections.find_one({"poi_id": poi_id})
    if not doc:
        return JsonResponse({"success": '', "msg": "该店暂无日志"})
    try:
        data = doc[date]
    except Exception as e:
        return JsonResponse({"success": '', "msg": "当天暂无日志"})
    return JsonResponse({"success": data, "msg": ''})


def set_tupian(request):
    if request.method == 'GET':
        arr = push_arr()
        return render(request, 'set_tupian.html', locals())

    poi_id = request.POST['poi_id']
    date = request.POST['date']
    text = request.POST['text']
    date_string = date
    date_format = "%Y-%m-%d"
    # 将日期字符串转换为datetime对象
    date_object = datetime.strptime(date_string, date_format)
    # 将datetime对象转换为时间戳
    timestamp = int(date_object.timestamp())

    client = MongoClient('mongodb://localhost:27017')
    other = client['other']
    collections = other['Essay']

    jtlog = client['jt_log']
    collections_log = jtlog[poi_id]

    doc = collections.find_one({"poi_id": poi_id})

    # print(text)
    # invalidorders = re.findall('今日刷单量：()', text)
    if "刷单量" not in text:
        return JsonResponse({"success": '', "msg": "该日志没有刷单量，请在第一行添加刷单量：？？？"})
    invalidorders = re.search(r'刷单量：(\d+)', text)
    if invalidorders:
        invalidorders = invalidorders.group(1)
    else:
        invalidorders = 0
    if invalidorders and invalidorders != '':
        temp = collections_log.find_one({"stime": timestamp})
        if temp:
            collections_log.update_one({"stime": timestamp}, {"$set": {"invalidOrders": invalidorders}})
        else:
            collections_log.insert_one({"ctime": date, "stime": timestamp, "invalidOrders": invalidorders})
    # 匹配“日志：”后面的所有文字
    log_match = re.search(r'日志：\n(.*)', text, re.DOTALL)
    if log_match:
        log_text = log_match.group(1).strip()
    else:
        log_text = ""
        return JsonResponse({"success": '保存，更新刷单量成功', "msg": ""})

    if log_text == '':
        return JsonResponse({"success": '保存，更新刷单量成功', "msg": ""})

    print("刷单量:", invalidorders, type(invalidorders))
    print("日志:", log_text)
    if not doc:
        collections.insert_one({"poi_id": poi_id, date: log_text})
        return JsonResponse({"success": '保存成功', "msg": ""})

    collections.update_one({"poi_id": poi_id}, {"$set": {date: log_text}})
    print(date, timestamp, poi_id, "保存日志成功")
    return JsonResponse({"success": '保存成功', "msg": ""})


# 夹心入库
def jx_rk(request, number, aa):
    if request.method == 'GET':
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'jx_rk.html', {"num": number, "aa": aa, "poi_id": poi_id, 'cookie': cookie, "arr": arr})
    poi_id = str(request.POST['poi_id'])
    cookie = request.POST['cookie'].strip()

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie

    num = 0
    if cache.has_key(request.POST['poi_id']):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id, 1)
    print('开始入库', poi_id)
    # set_database_jx01(poi_id, cookie)
    try:
        jx = JX(poi_id, cookie)
        jx.get_product()
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        return JsonResponse({'succ': '', 'msg': "数据更新失败，请联系管理员"})
    cache.delete(poi_id)

    return JsonResponse({'succ': '数据入库/更新数据成功', 'msg': ""})


# 折扣入库
def zk_rk(request, number, aa):
    if request.method == 'GET':
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'zk_rk.html', {"num": number, "aa": aa, "poi_id": poi_id, 'cookie': cookie, "arr": arr})
    poi_id = str(request.POST['poi_id'])

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id, 1)

    cookie = request.POST['cookie'].strip()
    result = "数据入库/更新数据成功"

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie

    print('开始入库', poi_id)
    try:
        try:
            result += set_database_jg01(poi_id, cookie)
        except Exception as e:
            print(e)
        set_database_jg02(poi_id, cookie)
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        return JsonResponse({'succ': '', 'msg': "数据更新失败，请联系管理员"})
    cache.delete(poi_id)
    return JsonResponse({'succ': result, 'msg': ""})
# 折扣更改

# 更改夹心
def updata_jx(request, number, aa):
    if request.method == 'GET':
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'updata_jx.html',
                      {"num": number, "aa": aa, "poi_id": poi_id, 'cookie': cookie, "arr": arr})
    poi_id = request.POST['poi_id']

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id, 1)

    # r = redis.Redis()
    # r.rpush('jx', poi_id)
    # while 1:
    #     pid = r.lindex('jx', 0).decode('utf-8')
    #     if pid == poi_id:
    #         break
    #     time.sleep(15)

    cookie = request.POST['cookie'].strip()

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie
    num = 0
    # 使用Django的文件存储系统保存文件
    # fs = FileSystemStorage()
    # filename = fs.save(file_obj.name, file_obj)
    # uploaded_file_url = fs.url(filename)
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        # r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        print(e)
        # 如果默认引擎失败，尝试使用 openpyxl（但请注意，这不会处理 .xls 文件）
        try:
            # 注意：这里我们不再使用 BytesIO，而是直接使用 file_obj，因为 pandas 支持文件对象
            df = pd.read_excel(file_obj, engine='openpyxl')
        except Exception as e:
            print(e)
            # 如果 openpyxl 也失败，则抛出错误或返回错误消息给用户
            cache.delete(poi_id)
            # r.lpop('jx')
            return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    reult = '更新成功，概率失败商品--》：'
    print('开始更改夹心', poi_id)
    set_arr = set()
    try:
        column_to_drop_duplicates_on = df.columns[2]
        df_no_duplicates = df.drop_duplicates(subset=column_to_drop_duplicates_on)
        # print(df_no_duplicates)


        jx = JX(poi_id, cookie)
        jx.updata_jx(df_no_duplicates, set_arr, 1)

        # asyncio.run(updata_Database_jx02(df_no_duplicates, cookie, poi_id, set_arr))
        # set_arr = set_jx(df_no_duplicates, cookie, poi_id)
        if set_arr == '0':
            raise '报错'
        # r.lpop('jx')
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        # r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': '更新失败，请重试/联系管理员', 'arr': list(set_arr)})
    # try:
    #     reult = updata_Database_jx02(df, cookie)
    # except Exception as e:
    #     print(e)
    print('完成夹心', poi_id)
    print(list(set_arr))
    cache.delete(poi_id)
    # fs.delete(file_obj.name)
    return JsonResponse({'succ': reult, 'msg': '', 'arr': list(set_arr)})


# 更改商品名
def updata_name(request, number, aa):
    if request.method == 'GET':
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'updata_jx.html',
                      {"num": number, "aa": aa, "poi_id": poi_id, 'cookie': cookie, "arr": arr})
    poi_id = request.POST['poi_id']

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id, 1)

    r = redis.Redis()
    r.rpush('jx', poi_id)
    while 1:
        pid = r.lindex('jx', 0).decode('utf-8')
        if pid == poi_id:
            break
        time.sleep(15)

    cookie = request.POST['cookie'].strip()

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie
    num = 0
    # 使用Django的文件存储系统保存文件
    # fs = FileSystemStorage()
    # filename = fs.save(file_obj.name, file_obj)
    # uploaded_file_url = fs.url(filename)
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        # 如果默认引擎失败，尝试使用 openpyxl（但请注意，这不会处理 .xls 文件）
        try:
            # 注意：这里我们不再使用 BytesIO，而是直接使用 file_obj，因为 pandas 支持文件对象
            df = pd.read_excel(file_obj, engine='openpyxl')
        except Exception as e:
            print(e)
            # 如果 openpyxl 也失败，则抛出错误或返回错误消息给用户
            cache.delete(poi_id)
            r.lpop('jx')
            return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    reult = '更新成功，概率失败商品--》：'
    print('开始更改夹心', poi_id)
    try:
        set_arr = rep_name(poi_id, df, cookie)
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': '更新失败，请重试/联系管理员'})
    # try:
    #     reult = updata_Database_jx02(df, cookie)
    # except Exception as e:
    #     print(e)
    print('完成夹心', poi_id)
    print(list(set_arr))
    cache.delete(poi_id)
    r.lpop('jx')
    # fs.delete(file_obj.name)
    return JsonResponse({'succ': reult, 'msg': '', 'arr': list(set_arr)})


# 更改折扣价格
def updata_jg(request, number, aa):
    if request.method == 'GET':
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'updata_jg.html',
                      {"num": number, "aa": aa, "poi_id": poi_id, 'cookie': cookie, "arr": arr})
    poi_id = request.POST['poi_id']

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    # cache.set(poi_id, 1)

    cookie = request.POST['cookie'].strip()

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie
    result = '更新成功，概率失败商品--》：'
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        print(e)
        # 如果默认引擎失败，尝试使用 openpyxl（但请注意，这不会处理 .xls 文件）
        try:
            # 注意：这里我们不再使用 BytesIO，而是直接使用 file_obj，因为 pandas 支持文件对象
            df = pd.read_excel(file_obj, engine='openpyxl')
        except Exception as e:
            print(e)
            # 如果 openpyxl 也失败，则抛出错误或返回错误消息给用户
            cache.delete(poi_id)
            return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    print('开始修改折扣', poi_id)
    onle = set()
    try:
        # onle = asyncio.run(updata_database_jg(df, poi_id, cookie))
        '''
        unsupported operand type(s) for -: 'int' and 'str'
        invalid literal for int() with base 10: '228..8'

        '''
        updata_database_jg(df, poi_id, cookie, onle)
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        return JsonResponse({'succ': '', 'msg': '请联系管理员', "arr": list(onle)})
    print('完成折扣', poi_id)
    print(list(onle))
    return JsonResponse({'succ': result, 'msg': '', "arr": list(onle)})


# 同步折扣价格
def sync_jg(request, number, aa):
    if request.method == 'GET':
        aa = aa
        poi_id_old = request.session.get('poi_id_old', '')
        poi_id_new = request.session.get('poi_id_old', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        num = number
        return render(request, 'sync_jg.html', locals())
    poi_id_old = request.POST['poi_id_old']
    poi_id_new = request.POST['poi_id_new']
    cookie = request.POST['cookie'].strip()

    # request.session['poi_id'] = poi_id
    # request.session['cookie'] = cookie

    if cache.has_key(poi_id_old):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id_old, 1)

    result = '更新成功，概率失败商品--》：'
    print('开始同步')
    # syncjg(poi_id_new, poi_id_old, cookie)
    setarr = set()
    try:
        syncjg(poi_id_new, poi_id_old, cookie, setarr)
    except Exception as e:
        print(e)
        cache.delete(poi_id_old)
        return JsonResponse({'succ': '', 'msg': "同步失败，请重新尝试，并联系管理员"})
    cache.delete(poi_id_old)
    return JsonResponse({'succ': result, 'msg': "", "arr": list(setarr)})


# 下载折扣表
def download_excel(request, poi_id):
    # 假设你有一个DataFrame
    client = MongoClient('mongodb://localhost:27017/')
    # 选择数据库和集合（相当于 SQL 中的表）
    db = client[str(poi_id)]
    collection = db['proact']
    documents = []
    dic = {}
    for i in collection.find():
        flog = 0
        dic[i['name']] = dic.get(i['name'], [])
        if 'errMsg' in i and i['errMsg'] == "已参与折扣活动":
            # pass
            dic[i['name']].append({
                '商品折扣id': i['actId'],
                '店铺id': poi_id,
                '分类名称': i['tagName'],
                '商品名字': i['name'],
                '商品规格': i['spec'],
                '原价': i['originPrice'],
                '折扣价': i['actPrice'],
            })
        # else:
        #     flog = 1
        #     dic[i['name']].append({
        #         '商品折扣id': None,
        #         '店铺id': poi_id,
        #         '分类名称': i['tagName'],
        #         '商品名字': i['name'],
        #         '商品规格': i['spec'],
        #         '原价': i['originPrice'],
        #         '折扣价': None,
        #     })

        # print(dic)
    for key, val in dic.items():
        val.sort(key=lambda x: x['原价'])
        # print(val)
        documents.extend(val)
    df = pd.DataFrame(documents)
    # 将DataFrame保存到Excel文件中
    excel_file = BytesIO()
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        # 将文件内容读回到内存
    excel_file.seek(0)

    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'

    return response


# 下载原价表
def download_excel_new(request, poi_id):
    """
    {
            '商品id': None,
            '店铺id': poi_id,
            '分类名称': i['tagName'],
            '商品名字': i['name'],
            '商品规格': i['spec'],
            '原价': i['originPrice'],
            '折扣价': None,

        }
    :param request:
    :param poi_id:
    :return:
    """
    # 假设你有一个DataFrame
    client = MongoClient('mongodb://localhost:27017/')
    # 选择数据库和集合（相当于 SQL 中的表）
    db = client[str(poi_id)]
    print(poi_id)
    collection = db["proact"]
    documents = []
    dic = {}
    for i in collection.find():
        dic[i['name']] = dic.get(i['name'], [])
        dic[i['name']].append({
            '商品折扣id': None,
            '店铺id': poi_id,
            '分类名称': i['tagName'],
            '商品名字': i['name'],
            '商品规格': i['spec'],
            '原价': i['originPrice'],
            '折扣价': None,
        })
    # print(dic)
    for key, val in dic.items():
        val.sort(key=lambda x: x['原价'])
        # print(val)
        documents.extend(val)

    # print(documents)
    df = pd.DataFrame(documents)
    # 将DataFrame保存到Excel文件中
    excel_file = BytesIO()
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        # 将文件内容读回到内存
    excel_file.seek(0)

    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'

    return response


# 下载夹心模板
def download_jx_excel(request):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动的工作表，默认是第一个工作表
    # 写入标题行
    headers = ['门店ID', '分类名称', '商品名字', '属性', '描述']
    ws.append(headers)
    # 定义字体样式为加粗
    header_font = Font(bold=True)

    # 定义对齐样式为居中
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取刚添加的标题行的索引（通常是最后一行）
    header_row = ws.max_row

    # 遍历标题行的每一个单元格，应用样式
    for cell in ws[header_row]:
        cell.font = header_font  # 加粗
        cell.alignment = header_alignment  # 居中
        cell.border = thin_border  # 添加边框

    excel_file = BytesIO()
    wb.save(excel_file)
    # 将BytesIO对象的指针移动到开始位置，以便从头读取内容
    excel_file.seek(0)
    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=jxTemp.xlsx'

    return response


# 下载折扣模板
def download_jg_excel(request):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动的工作表，默认是第一个工作表

    # 写入标题行
    headers = ['商品折扣id', '店铺id', '分类名称', '商品名字', '商品规格', '原价', '折扣价']
    ws.append(headers)

    # 定义字体样式为加粗
    header_font = Font(bold=True)

    # 定义对齐样式为居中
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取刚添加的标题行的索引（通常是最后一行）
    header_row = ws.max_row

    # 遍历标题行的每一个单元格，应用样式
    for cell in ws[header_row]:
        cell.font = header_font  # 加粗
        cell.alignment = header_alignment  # 居中
        cell.border = thin_border  # 添加边框

    excel_file = BytesIO()
    wb.save(excel_file)
    # 将BytesIO对象的指针移动到开始位置，以便从头读取内容
    excel_file.seek(0)
    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=jgTemp.xlsx'

    return response


# 下载精准营销模板
def download_jzyx_excel(request):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动的工作表，默认是第一个工作表

    # 写入标题行
    headers = ['cookie', '店铺id', '目标人群', '使用门槛', '优惠金额', '券有效期']
    ws.append(headers)

    # 定义字体样式为加粗
    header_font = Font(bold=True)

    # 定义对齐样式为居中
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取刚添加的标题行的索引（通常是最后一行）
    header_row = ws.max_row

    # 遍历标题行的每一个单元格，应用样式
    for cell in ws[header_row]:
        cell.font = header_font  # 加粗
        cell.alignment = header_alignment  # 居中
        cell.border = thin_border  # 添加边框

    excel_file = BytesIO()
    wb.save(excel_file)
    # 将BytesIO对象的指针移动到开始位置，以便从头读取内容
    excel_file.seek(0)
    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=jzyxTemp.xlsx'

    return response


# 下载月售替换模板
def download_replace_excel(request):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动的工作表，默认是第一个工作表

    # 写入标题行
    headers = ['产品名1', '折扣原名', '产品名2', '折扣原名2']
    ws.append(headers)

    # 定义字体样式为加粗
    header_font = Font(bold=True)

    # 定义对齐样式为居中
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取刚添加的标题行的索引（通常是最后一行）
    header_row = ws.max_row

    # 遍历标题行的每一个单元格，应用样式
    for cell in ws[header_row]:
        cell.font = header_font  # 加粗
        cell.alignment = header_alignment  # 居中
        cell.border = thin_border  # 添加边框

    excel_file = BytesIO()
    wb.save(excel_file)
    # 将BytesIO对象的指针移动到开始位置，以便从头读取内容
    excel_file.seek(0)
    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=replaceTemp.xlsx'

    return response


def gongzuoribao(request, username):
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动的工作表，默认是第一个工作表

    print(username)

    # 写入标题行
    headers = ['日期', '门店ID', '门店名称', '单量','营业额','推广费用','补单数（单量/金额）',
               "运营负责人", "平台"]
    ws.append(headers)

    # 定义字体样式为加粗
    header_font = Font(bold=True)

    # 定义对齐样式为居中
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 定义边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取刚添加的标题行的索引（通常是最后一行）
    header_row = ws.max_row

    # 遍历标题行的每一个单元格，应用样式
    for cell in ws[header_row]:
        cell.font = header_font  # 加粗
        cell.alignment = header_alignment  # 居中
        cell.border = thin_border  # 添加边框

    get_data(username, ws)

    excel_file = BytesIO()
    wb.save(excel_file)
    # 将BytesIO对象的指针移动到开始位置，以便从头读取内容
    excel_file.seek(0)
    # 设置HTTP响应头部
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={username}日报.xlsx'

    return response

# 下载折扣表请求页面
def dow_exc(request, number, aa):
    if request.method == "GET":
        num = number
        poi_id = request.session.get('poi_id', '')
        arr = push_arr()
        return render(request, 'download_excel.html', {"num": number, "aa": aa, "poi_id": poi_id, "arr": arr})

    poi_id = request.POST['poi_id']
    isnew = int(request.POST['is_new'])
    if isnew:
        url = f'/download_excel_new/{poi_id}'
    else:
        url = f'/dow_exc/{poi_id}'
    data = {'url': url}
    return JsonResponse(data)


# 搬菜单
def move_product(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id_old = request.session.get('poi_id_old', '')
        poi_id_new = request.session.get('poi_id_old', '')
        old_cookie = request.session.get('old_cookie', '')
        arr = push_arr()
        return render(request, 'move_product.html', locals())

    poi_id_old = request.POST['poi_id_old'].strip()
    poi_id_new = request.POST['poi_id_new'].strip()

    if cache.has_key(poi_id_old):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id_old, 1)

    r = redis.Redis()
    r.rpush('jx', poi_id_new)
    while 1:
        pid = r.lindex('jx', 0).decode('utf-8')
        if pid == poi_id_new:
            break
        time.sleep(15)

    cookie = request.POST['cookie'].strip()
    cookie_old = request.POST['cookie_old'].strip()
    description = request.POST['description'].strip()
    result = '同步完成, 同步失败商品: '
    # print(poi_id_old, poi_id_new)
    # move_pro(poi_id_old, poi_id_new, cookie, cookie_old)
    set_result = set()
    try:

        flog = move_pro(poi_id_old, poi_id_new, description, cookie, cookie_old, set_result)
        if flog == '0':
            raise ValueError("错误")
    except Exception as e:
        print(e)
        cache.delete(poi_id_old)
        r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': "同步失败，请重新尝试，并联系管理员", "arr": list(set_result)})
    cache.delete(poi_id_old)
    r.lpop('jx')
    return JsonResponse({'succ': result, 'msg': "", "arr": list(set_result)})


# 月售替换  redis.exceptions.ResponseError: syntax error
def replace_priduct(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'replace_priduct.html', locals())

    poi_id = request.POST['poi_id'].strip()
    name_x = request.POST['pro_new'].strip()
    name_y = request.POST['pro_old'].strip()
    cookie = request.POST['cookie'].strip()

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})

    cache.set(poi_id, 1)
    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie

    r = redis.Redis()
    temp = r.lindex('jx', 0)
    if temp:
        r.linsert('jx', 'after', temp.decode('utf-8'), poi_id)
    else:
        r.rpush('jx', poi_id)

    while 1:
        pid = r.lindex('jx', 0).decode('utf-8')
        if pid == poi_id:
            break

        time.sleep(15)
    print(f'开始替换商品：{poi_id, name_x, name_y}')
    result = set()
    try:
        jx = JX(poi_id, cookie)
        jx.replace_product(name_x, name_y, result)
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', 'list': list(result)})

    if len(result):
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', "list": list(result)})
    reult = {"succ": '替换成功', 'msg': None}
    cache.delete(poi_id)
    r.lpop('jx')
    return JsonResponse(reult)


# 月售替换
def replace_priduct1(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'replace_priduct2.html', locals())

    poi_id = request.POST['poi_id'].strip()
    name_x = request.POST['pro_new'].strip()
    name_y = request.POST['pro_old'].strip()
    act_name_x = request.POST['pro_new_act'].strip()
    act_name_y = request.POST['pro_old_act'].strip()
    cookie = request.POST['cookie'].strip()

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})

    cache.set(poi_id, 1)
    r = redis.Redis()
    temp = r.lindex('jx', 0)
    if temp:
        r.linsert('jx', 'after', temp.decode('utf-8'), poi_id)
    else:
        r.rpush('jx', poi_id)

    while 1:
        pid = r.lindex('jx', 0).decode('utf-8')
        if pid == poi_id:
            break
        time.sleep(15)

    print(f'开始替换商品：{poi_id, name_x, name_y}')

    result = set()
    try:
        jx = JX(poi_id, cookie)
        jx.replace_product(name_x, name_y, result)
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', 'list': list(result)})

    if len(result):
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', "list": list(result)})
    reult = {"succ": '替换成功', 'msg': None}
    cache.delete(poi_id)
    r.lpop('jx')
    return JsonResponse(reult)
    # return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员'})  'NoneType' object is not subscriptable


# 月售替换
def replace_priduct2(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'replace_priduct3.html', locals())
    poi_id = request.POST['poi_id']
    cookie = request.POST['cookie'].strip()

    if cache.has_key(poi_id):
        return JsonResponse({'succ': '', 'msg': "该店正在操作，请等待结束"})

    cache.set(poi_id, 1)
    r = redis.Redis()
    temp = r.lindex('jx', 0)
    if temp:
        r.linsert('jx', 'after', temp.decode('utf-8'), poi_id)
    else:
        r.rpush('jx', poi_id)

    while 1:
        pid = r.lindex('jx', 0).decode('utf-8')
        if pid == poi_id:
            break
        time.sleep(15)
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        # 如果默认引擎失败，尝试使用 openpyxl（但请注意，这不会处理 .xls 文件）
        try:
            # 注意：这里我们不再使用 BytesIO，而是直接使用 file_obj，因为 pandas 支持文件对象
            df = pd.read_excel(file_obj, engine='openpyxl')
        except Exception as e:
            # 如果 openpyxl 也失败，则抛出错误或返回错误消息给用户
            cache.delete(poi_id)
            r.lpop('jx')
            return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    # print(f'开始替换商品：{poi_id}')
    result = set()
    try:
        rep_pro3(poi_id, df, cookie, result)
    except Exception:
        cache.delete(poi_id)
        r.lpop('jx')
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', "list": list(result)})

    if len(result):
        cache.delete(poi_id)
        return JsonResponse({"succ": None, 'msg': '替换失败，请联系管理员', "list": list(result)})

    reult = {"succ": '替换成功', 'msg': None}
    cache.delete(poi_id)
    r.lpop('jx')
    return JsonResponse(reult)


# 　好评回复
def review_reply(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        arr = push_arr()
        return render(request, 'review_reply.html', locals())

    poi_id = request.POST['poi_id']
    cookie = request.POST['cookie']
    content = request.POST['content']

    print(poi_id, "好评回复")

    client = MongoClient('mongodb://localhost:27017/')
    db = client['reply']
    collect = db['all']

    doc = collect.find_one({"wmpoid": poi_id})

    if doc:
        collect.update_one({"wmpoid": poi_id}, {"$set": {
            "wmpoid": poi_id,
            "content": content
        }})
    else:
        collect.insert_one({
            "wmpoid": poi_id,
            "content": content
        })

    result = {"succ": "回复完成", 'msg': ""}
    try:
        auto_rep(poi_id, cookie)
    except Exception as e:
        print(e)
        JsonResponse({"succ": "", "msg": "设置失败，请联系管理员"})
    return JsonResponse(result)


# 定时修改
def ds(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        cookie = request.session.get('cookie', '')
        hourses = range(24)
        arr = push_arr()
        return render(request, 'updata_jx_ds.html', locals())

    poi_id = request.POST['poi_id']

    cookie = request.POST['cookie'].strip()
    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie

    type_select = request.POST['type_select']
    date = request.POST['date']
    datetime_obj = datetime.fromisoformat(date)

    print(f"{datetime_obj.year}_{datetime_obj.month}_{datetime_obj.day}_{datetime_obj.hour}_{datetime_obj.minute}")
    save_path = r'G:/updata/定时修改/'
    backup_file_path = f'G:/backup/{type_select}/'
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        return JsonResponse({"succ": "", "msg": "请上传文件"})
    extension = os.path.splitext(file_obj.name)[1]
    file_name = os.path.splitext(file_obj.name)[0]
    make_an_appointment = f"{datetime_obj.year}_{datetime_obj.month}_{datetime_obj.day}_{datetime_obj.hour}_{datetime_obj.minute}"
    new_file_name = f"{file_name}_{make_an_appointment}_{type_select}{extension}"
    backup_file_name = f"{file_name}_{make_an_appointment}_{type_select}{extension}"

    # 在指定目录下查找是否有同名文件
    existing_files = [f for f in os.listdir(save_path) if
                      f.startswith(new_file_name) and f.endswith(extension)]
    if existing_files:
        # 如果存在同名文件，删除它以便于覆盖保存
        os.remove(os.path.join(save_path, existing_files[0]))

    # 使用Django的文件存储系统保存文件
    fs = FileSystemStorage(location=save_path)
    backup_fs = FileSystemStorage(location=backup_file_path)
    filename = fs.save(new_file_name, file_obj)
    backup_fs.save(backup_file_name, file_obj)

    uploaded_file_url = save_path + filename

    save_ds(poi_id, cookie, datetime_obj, type_select, new_file_name, uploaded_file_url)
    return JsonResponse({'success': '成功', 'msg': '', 'arr': ''})


# 查看定时修改任务
def get_ds(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        arr = push_arr()
        if poi_id:
            docs = push_ds(poi_id)
        return render(request, 'get_ds.html', locals())
    poi_id = request.POST['poi_id']
    request.session['poi_id'] = poi_id
    return JsonResponse({'success': '成功'})


# 删除定时修改任务
def delds(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        poi_id = request.session.get('poi_id', '')
        arr = push_arr()
        if poi_id:
            docs = push_ds(poi_id)
        return render(request, 'get_ds.html', locals())

    dsid = request.POST['dsid']
    print(dsid)
    delds_rw(dsid)
    return JsonResponse({"success": "成功"})


# 设置自动 精准营销
def jzyx_2(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        return render(request, 'auto_cup.html', locals())

    poi_id = request.POST['poi_id']
    limit_price = request.POST['limit_price']
    price = request.POST['price']
    day = request.POST['day']
    cookie = request.POST['cookie']
    type_select = request.POST['type_select']
    time_select = request.POST['time_select']

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie

    # print(poi_id)
    # print(limit_price)
    # print(price)
    # print(day)
    # print(cookie)
    # print(type_select)

    save_jz(poi_id, limit_price, price, day, cookie, type_select, time_select)

    return JsonResponse({"success": "成功", "msg": ""})


# 手动精准营销
def jzyx_1(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        return render(request, '精准营销.html', locals())
    try:
        file_obj = request.FILES['file_obj']
    except Exception as e:
        print(e)
        return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        # 如果默认引擎失败，尝试使用 openpyxl（但请注意，这不会处理 .xls 文件）
        try:
            # 注意：这里我们不再使用 BytesIO，而是直接使用 file_obj，因为 pandas 支持文件对象
            df = pd.read_excel(file_obj, engine='openpyxl')
        except Exception as e:
            # 如果 openpyxl 也失败，则抛出错误或返回错误消息给用户
            return JsonResponse({'succ': '', 'msg': "文件格式错误，请核对文件"})
    try:
        result = yx(df)
        if len(result):
            return JsonResponse({"succ": "", 'msg': "发送失败， 请联系管理员", 'arr': list(result)})
    except Exception as e:
        print(e)
        return JsonResponse({"succ": "", 'msg': "发送失败,请联系管理员"})
    return JsonResponse({"succ": "发送成功", 'msg': ""})


# 店铺数据备份
def back_poi(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        try:
            poi_id = request.session['poi_id']
        except Exception as e:
            return render(request, 'back_wmpoi.html', {"poi_id": '', "num": num, "aa": aa})
        return render(request, 'back_wmpoi.html', locals())

    poi_id = request.POST['poi_id']
    request.session['poi_id'] = poi_id
    back(poi_id)
    return JsonResponse({'succ': "成功", "msg": ""})


# 店铺折扣价格数备份
def back_jgpoi(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        poi_id = request.session['poi_id']
        return render(request, 'back_wmpoi.html', locals())

    poi_id = request.POST['poi_id']
    request.session['poi_id'] = poi_id
    back_zk(poi_id)
    print(poi_id, "备份折扣成功")
    return JsonResponse({'succ': "成功", "msg": ""})


# 查看数据备份
def get_back_poi(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        poi_id = request.session.get('poi_id')
        cookie = request.session.get('cookie')
        documents = get_back_log(poi_id)
        return render(request, 'get_back_poi.html', locals())
    poi_id = request.POST['poi_id']
    # get_back_log(poi_id)
    request.session['poi_id'] = poi_id
    return JsonResponse({'success': "成功", "msg": ""})


# 备份数据恢复
def back_Recover(request, number, aa):
    if request.method == 'GET':
        num = number
        aa = aa
        arr = push_arr()
        try:
            poi_id = request.session['poi_id']
            cookie = request.session['cookie']
        except Exception as e:
            return render(request, 'back_wmpoi.html', {"poi_id": '', "num": num, 'aa': aa})
        return render(request, 'back_wmpoi.html', locals())

    type_select = request.POST['type_select']
    back_poid = request.POST['back_poid']
    cookie = request.POST['cookie']
    poi_id = request.POST['poi_id']

    request.session['poi_id'] = poi_id
    request.session['cookie'] = cookie
    if cache.has_key(request.POST['poi_id']):
        return JsonResponse({'success': '', 'msg': "该店正在操作，请等待结束"})
    cache.set(poi_id, 1)

    try:
        if type_select == "0":
            r = redis.Redis()
            r.rpush('jx', poi_id)
            while 1:
                pid = r.lindex('jx', 0).decode('utf-8')
                if pid == poi_id:
                    break
                time.sleep(15)
            try:
                jx = JX(poi_id, cookie)
                jx.get_product()
                jx.run_Recover(back_poid, 0)
                # set_database_jx01(poi_id, cookie)
                # asyncio.run(Recover(back_poid, poi_id, cookie))
                r.lpop("jx")
            except Exception as e:
                print(e)
                r.lpop("jx")
        else:
            set_database_jg01(poi_id, cookie)
            set_database_jg02(poi_id, cookie)
            Recover_jg(back_poid, poi_id, cookie)
        cache.delete(poi_id)


    except Exception as e:
        cache.delete(poi_id)
        print(e)
    return JsonResponse({"success": '恢复成功'})


"""
asyncio.run(main(df))

'utf-16-le' codec can't decode bytes in position 20-21: unexpected end of data
File is not a zip file

redis.exceptions.ResponseError: syntax error

"""
